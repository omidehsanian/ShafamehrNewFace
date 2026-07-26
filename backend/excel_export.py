# -*- coding: utf-8 -*-
"""
ماژول خروجی گرفتن اکسل از اطلاعات بیماران، نوبت‌ها و پیشرفت درمان
از openpyxl استفاده می‌شود.
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2E5C8A", end_color="2E5C8A", fill_type="solid")
BODY_FONT = Font(name="Arial", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)


def _style_sheet(ws, headers, col_widths, rtl=True):
    ws.sheet_view.rightToLeft = rtl
    ws.freeze_panes = "A2"
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths[col_idx - 1]
    ws.row_dimensions[1].height = 24


def _write_row(ws, row_idx, values):
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value if value not in (None, "") else "")
        cell.font = BODY_FONT
        cell.alignment = RIGHT
        cell.border = THIN_BORDER


def export_patients(patients, filepath):
    """
    patients: لیستی از sqlite3.Row (خروجی db.get_patients())
    filepath: مسیر کامل فایل خروجی .xlsx
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "بیماران"

    headers = [
        "نام", "نام خانوادگی", "کد ملی", "شماره تلفن", "جنسیت",
        "تاریخ تولد", "آدرس", "شغل",
        "بیمه پایه", "کد رهگیری نسخه", "بیمه تکمیلی", "نام سازمان (بیمه تکمیلی)",
        "تشخیص / نوع بیماری", "سابقه پزشکی", "حساسیت دارویی", "پزشک معالج",
        "تماس اضطراری (نام)", "تماس اضطراری (تلفن)",
        "گروه خونی", "قد (cm)", "وزن (kg)", "یادداشت", "تاریخ ثبت",
    ]
    col_widths = [14, 16, 14, 15, 8, 15, 22, 14, 22, 16, 18, 16, 22, 22,
                  18, 16, 16, 16, 10, 8, 8, 25, 16]

    _style_sheet(ws, headers, col_widths)

    for i, p in enumerate(patients, start=2):
        _write_row(ws, i, [
            p["first_name"], p["last_name"], p["national_code"], p["phone"], p["gender"],
            p["birth_date"], p["address"], p["occupation"],
            p["primary_insurance"], p["prescription_code"],
            p["supplementary_insurance"], p["supplementary_insurance_org"],
            p["diagnosis"], p["medical_history"], p["allergies"], p["referring_doctor"],
            p["emergency_contact_name"], p["emergency_contact_phone"],
            p["blood_type"], p["height"], p["weight"], p["notes"], p["created_at"],
        ])

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(len(patients) + 1, 1)}"
    wb.save(filepath)
    return filepath


def export_appointments(appointments, filepath):
    """appointments: خروجی db.get_appointments()"""
    wb = Workbook()
    ws = wb.active
    ws.title = "نوبت‌ها"

    headers = ["نام بیمار", "تاریخ", "ساعت", "وضعیت", "یادداشت"]
    col_widths = [20, 14, 10, 16, 30]
    _style_sheet(ws, headers, col_widths)

    for i, a in enumerate(appointments, start=2):
        _write_row(ws, i, [
            a["full_name"], a["appointment_date"], a["appointment_time"],
            a["status"], a["notes"],
        ])

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(len(appointments) + 1, 1)}"
    wb.save(filepath)
    return filepath


def export_patient_full_profile(patient, appointments, exercises, progress_logs, filepath):
    """
    یک خروجی اکسل چند شیتی برای پرونده کامل یک بیمار خاص:
    شیت اول مشخصات، شیت دوم نوبت‌ها، شیت سوم تمرینات، شیت چهارم روند پیشرفت
    """
    wb = Workbook()

    # ---------- شیت مشخصات ----------
    ws1 = wb.active
    ws1.title = "مشخصات بیمار"
    ws1.sheet_view.rightToLeft = True
    ws1.column_dimensions["A"].width = 26
    ws1.column_dimensions["B"].width = 34

    fields = [
        ("نام", patient["first_name"]), ("نام خانوادگی", patient["last_name"]),
        ("کد ملی", patient["national_code"]), ("شماره تلفن", patient["phone"]),
        ("جنسیت", patient["gender"]), ("تاریخ تولد", patient["birth_date"]),
        ("آدرس", patient["address"]), ("شغل", patient["occupation"]),
        ("بیمه پایه", patient["primary_insurance"]), ("کد رهگیری نسخه", patient["prescription_code"]),
        ("بیمه تکمیلی", patient["supplementary_insurance"]),
        ("نام سازمان (بیمه تکمیلی)", patient["supplementary_insurance_org"]),
        ("تشخیص / نوع بیماری", patient["diagnosis"]), ("سابقه پزشکی", patient["medical_history"]),
        ("حساسیت دارویی", patient["allergies"]), ("پزشک معالج", patient["referring_doctor"]),
        ("تماس اضطراری (نام)", patient["emergency_contact_name"]),
        ("تماس اضطراری (تلفن)", patient["emergency_contact_phone"]),
        ("گروه خونی", patient["blood_type"]), ("قد (cm)", patient["height"]),
        ("وزن (kg)", patient["weight"]), ("یادداشت", patient["notes"]),
        ("تاریخ ثبت", patient["created_at"]),
    ]
    for i, (label, value) in enumerate(fields, start=1):
        c1 = ws1.cell(row=i, column=1, value=label)
        c1.font = Font(name="Arial", bold=True, size=10)
        c1.alignment = RIGHT
        c1.fill = PatternFill(start_color="EAF0F8", end_color="EAF0F8", fill_type="solid")
        c2 = ws1.cell(row=i, column=2, value=value if value not in (None, "") else "")
        c2.font = BODY_FONT
        c2.alignment = RIGHT

    # ---------- شیت نوبت‌ها ----------
    ws2 = wb.create_sheet("نوبت‌ها")
    headers2 = ["تاریخ", "ساعت", "وضعیت", "یادداشت"]
    _style_sheet(ws2, headers2, [14, 10, 16, 30])
    for i, a in enumerate(appointments, start=2):
        _write_row(ws2, i, [a["appointment_date"], a["appointment_time"], a["status"], a["notes"]])

    # ---------- شیت تمرینات ----------
    ws3 = wb.create_sheet("تمرینات")
    headers3 = ["نام تمرین", "دسته", "ست", "تکرار", "زمان (ثانیه)", "تاریخ تخصیص"]
    _style_sheet(ws3, headers3, [20, 14, 8, 8, 12, 14])
    for i, e in enumerate(exercises, start=2):
        _write_row(ws3, i, [e["name"], e["category"], e["sets"], e["reps"], e["duration_sec"], e["assigned_date"]])

    # ---------- شیت روند پیشرفت ----------
    ws4 = wb.create_sheet("روند پیشرفت")
    headers4 = ["تاریخ", "میزان درد (۰-۱۰)", "دامنه حرکتی (۰-۱۰۰)", "یادداشت"]
    _style_sheet(ws4, headers4, [14, 16, 18, 30])
    for i, l in enumerate(progress_logs, start=2):
        _write_row(ws4, i, [l["log_date"], l["pain_level"], l["mobility_score"], l["notes"]])

    wb.save(filepath)
    return filepath


# ---------------- ایمپورت دسته‌ای بیماران از اکسل ----------------

# ترتیب و عنوان ستون‌های قالب ایمپورت. کلید سمت راست، نام ستون در دیتابیس است.
IMPORT_COLUMNS = [
    ("نام *", "first_name"),
    ("نام خانوادگی *", "last_name"),
    ("کد ملی", "national_code"),
    ("شماره تلفن", "phone"),
    ("جنسیت (مرد/زن)", "gender"),
    ("تاریخ تولد (شمسی، مثال 1370/05/12)", "birth_date"),
    ("بیمه پایه", "primary_insurance"),
    ("بیمه تکمیلی", "supplementary_insurance"),
    ("تشخیص / نوع بیماری", "diagnosis"),
]


def export_patient_import_template(filepath):
    """یک فایل اکسل خالی (فقط با هدر) برای بارگذاری دسته‌ای بیماران می‌سازد"""
    wb = Workbook()
    ws = wb.active
    ws.title = "قالب بیماران"

    headers = [col[0] for col in IMPORT_COLUMNS]
    col_widths = [16, 18, 14, 15, 16, 26, 20, 20, 22]
    _style_sheet(ws, headers, col_widths)

    # یک ردیف نمونه برای راهنمایی کاربر (بعداً باید پاک یا جایگزین شود)
    _write_row(ws, 2, ["علی", "رضایی", "1234567890", "09121234567", "مرد",
                        "1370/05/12", "تأمین اجتماعی", "بیمه دانا", "کمردرد مزمن"])

    wb.save(filepath)
    return filepath


def import_patients_from_excel(filepath, db):
    """
    فایل اکسل بارگذاری‌شده را می‌خواند و بیماران معتبر را به دیتابیس اضافه می‌کند.
    خروجی: (تعداد موفق, لیست خطاها به‌صورت رشته)
    """
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    keys = [col[1] for col in IMPORT_COLUMNS]
    success_count = 0
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(cell in (None, "") for cell in row):
            continue  # ردیف کاملاً خالی را رد کن

        data = {}
        for i, key in enumerate(keys):
            value = row[i] if i < len(row) else None
            data[key] = str(value).strip() if value not in (None, "") else ""

        if not data.get("first_name") or not data.get("last_name"):
            errors.append(f"ردیف {row_idx}: نام یا نام خانوادگی خالی است — رد شد.")
            continue

        try:
            db.add_patient(data)
            success_count += 1
        except Exception as e:
            errors.append(f"ردیف {row_idx}: خطا در ثبت — {e}")

    return success_count, errors
